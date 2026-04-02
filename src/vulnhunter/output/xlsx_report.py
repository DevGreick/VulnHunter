import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from vulnhunter.models import ScanResult, Severity

logger = logging.getLogger(__name__)

SEVERITY_COLORS: dict[str, str] = {
    "CRITICAL": "FF4444",
    "HIGH": "FF8C00",
    "MEDIUM": "FFD700",
    "LOW": "27C93F",
    "UNKNOWN": "8B949E",
}

HEADER_FILL = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="00E5FF", size=11)
HEADER_BORDER = Border(bottom=Side(style="thin", color="1A2332"))
CELL_FONT = Font(name="Calibri", size=10, color="E6EDF3")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="00E5FF")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="8B949E")


def _validate_output_path(output_path: Path, base_dir: Path | None = None) -> Path:
    resolved = output_path.resolve()
    if base_dir is not None:
        base_resolved = base_dir.resolve()
        if not resolved.is_relative_to(base_resolved):
            raise ValueError("output path escapes base directory")
    return resolved


def _style_header(ws: Worksheet, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _severity_fill(sev: str) -> PatternFill:
    color = SEVERITY_COLORS.get(sev, "8B949E")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _severity_font(sev: str) -> Font:
    bg = SEVERITY_COLORS.get(sev, "8B949E")
    text_color = "000000" if sev in ("MEDIUM", "LOW") else "FFFFFF"
    return Font(name="Calibri", bold=True, size=10, color=text_color)


def _build_summary_sheet(wb: Workbook, result: ScanResult) -> None:
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="VulnHunter Scan Report").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"VulnHunter v2.0.0").font = SUBTITLE_FONT

    ws.cell(row=5, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=5, column=1).fill = HEADER_FILL
    ws.cell(row=5, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=5, column=2).fill = HEADER_FILL

    metrics: list[tuple[str, Any]] = [
        ("Dependencies Scanned", result.total_dependencies),
        ("Vulnerabilities Found", result.total_vulnerabilities),
        ("Ignored", result.total_ignored),
    ]

    severity_counts: dict[str, int] = {}
    for vuln in result.vulnerabilities:
        sev = vuln.severity.value
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    row = 6
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label).font = CELL_FONT
        ws.cell(row=row, column=2, value=value).font = Font(name="Calibri", bold=True, size=10, color="00E5FF")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Severity Breakdown").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1

    for sev in [Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM, Severity.LOW, Severity.UNKNOWN]:
        count = severity_counts.get(sev.value, 0)
        cell_sev = ws.cell(row=row, column=1, value=sev.value)
        cell_sev.fill = _severity_fill(sev.value)
        cell_sev.font = _severity_font(sev.value)
        cell_sev.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=count).font = Font(name="Calibri", bold=True, size=10, color="E6EDF3")
        row += 1

    ecosystems_seen: set[str] = set()
    for vuln in result.vulnerabilities:
        ecosystems_seen.add(vuln.ecosystem.value)

    row += 1
    ws.cell(row=row, column=1, value="Ecosystems Scanned").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for eco in sorted(ecosystems_seen):
        ws.cell(row=row, column=1, value=eco).font = CELL_FONT
        row += 1

    _auto_width(ws)


def _build_vulnerabilities_sheet(wb: Workbook, result: ScanResult) -> None:
    ws = wb.create_sheet(title="Vulnerabilities")

    headers = ["CVE ID", "Severity", "Package", "Version", "Fixed", "Ecosystem", "Source", "Summary"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header(ws, len(headers))

    for row_idx, vuln in enumerate(result.vulnerabilities, 2):
        ws.cell(row=row_idx, column=1, value=vuln.vuln_id).font = CELL_FONT
        sev_cell = ws.cell(row=row_idx, column=2, value=vuln.severity.value)
        sev_cell.fill = _severity_fill(vuln.severity.value)
        sev_cell.font = _severity_font(vuln.severity.value)
        sev_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=vuln.name).font = CELL_FONT
        ws.cell(row=row_idx, column=4, value=vuln.version).font = CELL_FONT
        ws.cell(row=row_idx, column=5, value=vuln.fixed_version or "N/A").font = Font(
            name="Calibri", size=10, color="27C93F" if vuln.fixed_version else "8B949E"
        )
        ws.cell(row=row_idx, column=6, value=vuln.ecosystem.value).font = CELL_FONT
        ws.cell(row=row_idx, column=7, value=vuln.source).font = CELL_FONT
        ws.cell(row=row_idx, column=8, value=vuln.summary).font = CELL_FONT

    ws.auto_filter.ref = f"A1:H{len(result.vulnerabilities) + 1}"
    _auto_width(ws)


def _build_ecosystem_sheet(wb: Workbook, result: ScanResult) -> None:
    eco_groups: dict[str, list[Any]] = {}
    for vuln in result.vulnerabilities:
        eco = vuln.ecosystem.value
        if eco not in eco_groups:
            eco_groups[eco] = []
        eco_groups[eco].append(vuln)

    ws = wb.create_sheet(title="By Ecosystem")

    headers = ["Ecosystem", "CVE ID", "Severity", "Package", "Version", "Fixed"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header(ws, len(headers))

    row = 2
    for eco in sorted(eco_groups.keys()):
        vulns = eco_groups[eco]
        for vuln in vulns:
            ws.cell(row=row, column=1, value=eco).font = CELL_FONT
            ws.cell(row=row, column=2, value=vuln.vuln_id).font = CELL_FONT
            sev_cell = ws.cell(row=row, column=3, value=vuln.severity.value)
            sev_cell.fill = _severity_fill(vuln.severity.value)
            sev_cell.font = _severity_font(vuln.severity.value)
            sev_cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=vuln.name).font = CELL_FONT
            ws.cell(row=row, column=5, value=vuln.version).font = CELL_FONT
            ws.cell(row=row, column=6, value=vuln.fixed_version or "N/A").font = Font(
                name="Calibri", size=10, color="27C93F" if vuln.fixed_version else "8B949E"
            )
            row += 1

    ws.auto_filter.ref = f"A1:F{row - 1}"
    _auto_width(ws)


def render_xlsx(result: ScanResult, output_path: Path, base_dir: Path | None = None) -> None:
    safe_path = _validate_output_path(output_path, base_dir)
    safe_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _build_summary_sheet(wb, result)
    _build_vulnerabilities_sheet(wb, result)
    _build_ecosystem_sheet(wb, result)
    wb.save(str(safe_path))
    logger.info("XLSX report written to %s", safe_path)
