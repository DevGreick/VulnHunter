from pathlib import Path

import pytest
from openpyxl import load_workbook

from vulnhunter.models import Dependency, Ecosystem, ScanResult, Severity, Vulnerability
from vulnhunter.output.xlsx_report import render_xlsx


@pytest.fixture
def sample_result() -> ScanResult:
    return ScanResult(
        total_dependencies=3,
        total_vulnerabilities=2,
        total_ignored=1,
        vulnerabilities=[
            Vulnerability(
                vuln_id="CVE-2023-0001",
                source="osv",
                name="flask",
                version="2.0.1",
                ecosystem=Ecosystem.PYPI,
                severity=Severity.HIGH,
                summary="Test vulnerability in flask",
                fixed_version="2.3.2",
            ),
            Vulnerability(
                vuln_id="CVE-2023-0002",
                source="nvd",
                name="express",
                version="4.17.1",
                ecosystem=Ecosystem.NPM,
                severity=Severity.CRITICAL,
                summary="Critical issue in express",
                fixed_version="4.18.0",
            ),
        ],
        dependencies=[
            Dependency(name="flask", version="2.0.1", ecosystem=Ecosystem.PYPI),
            Dependency(name="requests", version="2.31.0", ecosystem=Ecosystem.PYPI),
            Dependency(name="express", version="4.17.1", ecosystem=Ecosystem.NPM),
        ],
    )


@pytest.fixture
def empty_result() -> ScanResult:
    return ScanResult()


def test_creates_valid_xlsx(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    assert output.exists()
    assert output.stat().st_size > 0


def test_has_three_sheets(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    wb = load_workbook(str(output))
    assert len(wb.sheetnames) == 3
    assert "Summary" in wb.sheetnames
    assert "Vulnerabilities" in wb.sheetnames
    assert "By Ecosystem" in wb.sheetnames
    wb.close()


def test_summary_sheet_has_correct_counts(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    wb = load_workbook(str(output))
    ws = wb["Summary"]
    values: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                values.append(str(cell))
    assert "3" in values
    assert "2" in values
    assert "1" in values
    wb.close()


def test_vulnerabilities_sheet_rows_match(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    wb = load_workbook(str(output))
    ws = wb["Vulnerabilities"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2
    cve_ids = {row[0] for row in data_rows}
    assert "CVE-2023-0001" in cve_ids
    assert "CVE-2023-0002" in cve_ids
    wb.close()


def test_vulnerabilities_sheet_headers(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    wb = load_workbook(str(output))
    ws = wb["Vulnerabilities"]
    headers = [cell.value for cell in ws[1]]
    assert "CVE ID" in headers
    assert "Severity" in headers
    assert "Package" in headers
    wb.close()


def test_ecosystem_sheet_groups_correctly(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    wb = load_workbook(str(output))
    ws = wb["By Ecosystem"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    ecosystems = {row[0] for row in data_rows if row[0]}
    assert "PyPI" in ecosystems
    assert "npm" in ecosystems
    wb.close()


def test_empty_result_creates_valid_xlsx(tmp_path: Path, empty_result: ScanResult) -> None:
    output = tmp_path / "report.xlsx"
    render_xlsx(empty_result, output, base_dir=tmp_path)
    assert output.exists()
    wb = load_workbook(str(output))
    assert len(wb.sheetnames) == 3
    ws = wb["Vulnerabilities"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 0
    wb.close()


def test_path_traversal_blocked(tmp_path: Path, sample_result: ScanResult) -> None:
    evil_path = tmp_path / ".." / ".." / "etc" / "evil.xlsx"
    with pytest.raises(ValueError, match="escapes base directory"):
        render_xlsx(sample_result, evil_path, base_dir=tmp_path)


def test_creates_parent_directories(tmp_path: Path, sample_result: ScanResult) -> None:
    output = tmp_path / "subdir" / "deep" / "report.xlsx"
    render_xlsx(sample_result, output, base_dir=tmp_path)
    assert output.exists()
